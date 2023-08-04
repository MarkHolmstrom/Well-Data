#!/usr/bin/env python3


import importlib
import os
import sys

import lasio
import numpy as np
from openpyxl import Workbook, drawing
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from PyQt5.QtWidgets import QApplication, QFileDialog


def modify_and_import(module_name, package, modification_func):
    """Function to modify module source code from https://stackoverflow.com/a/41863728"""
    spec = importlib.util.find_spec(module_name, package)
    source = spec.loader.get_source(module_name)
    new_source = modification_func(source)
    module = importlib.util.module_from_spec(spec)
    codeobj = compile(new_source, module.__spec__.origin, 'exec')
    exec(codeobj, module.__dict__)
    sys.modules[module_name] = module
    return module


def missing_dt_fill(i, dept, prev_dept, formation_tvd_array):
    """Function for efficiently generating cell references that do not cross formation boundaries."""
    if any(prev_dept < formation_bound <= dept for formation_bound in formation_tvd_array):  # boundary cross
        return "NaN", formation_tvd_array[formation_tvd_array > dept]  # remove crossed boundaries
    else:
        return str(i - 1), formation_tvd_array  # no boundary cross, return reference to prev measurement


def make_excel_manual(info, las_df, coefficient_df):
    """Given a dictionary of well information and a DataFrame containing las depth information,
        return an Excel workbook formatted with it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "basic infor"
    ws.sheet_view.zoomScale = 124
    # zoom does not work as mentioned in https://openpyxl.readthedocs.io/en/stable/worksheet_properties.html
    width_const = 0.77
    ws.column_dimensions["A"].width = 18.33 + width_const
    ws.column_dimensions["B"].width = 8.11 + width_const
    ws.column_dimensions["C"].width = 5.78 + width_const
    ws.column_dimensions["D"].width = 8.33 + width_const
    ws.column_dimensions["E"].width = 8.89 + width_const
    ws.column_dimensions["F"].width = 8.11 + width_const
    ws.column_dimensions["G"].width = 22.56 + width_const
    ws.column_dimensions["H"].width = 10.78 + width_const
    ws.column_dimensions["I"].width = 9.89 + width_const
    ws.column_dimensions["J"].width = 13.67 + width_const
    # match given column widths

    emu_to_pt_const = 10000 * 1.27

    ws["A3"], ws["D3"], ws["G3"], ws["G4"] = "Lati", "Long", "Strain x", "Strain y"
    ws["A5"], ws["D5"], ws["A7"], ws["A9"] = "KB (m):", "GL (m):", "Mud weight (kg/m^3)", "Formation tops"
    ws["B9"], ws["C9"], ws["D9"], ws["B10"], ws["C10"], ws["D10"] = "TVD", "Elev", "MD", "(m)", "(m)", "(m)"
    ws["G10"], ws["G11"] = "100/06-03-079-13W6_00", "100/13-35-081-21W6_00"
    ws["H9"], ws["I9"], ws["J9"] = "Long", "strain x", "strain y"
    ws["H10"], ws["H11"], ws["I10"], ws["I11"], ws["J10"] = -119.94717, -121.18236, 0.02, 0.04, 0.26
    ws["J11"], ws["I12"], ws["J12"] = 0.52, "=(I11-I10)/(H10-H11)", "=(J11-J10)/(H10-H11)"
    # static strings

    for row_count, row in enumerate(ws["G9":"J12"]):
        for col_count, cell in enumerate(row):
            cell.fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")
    # strain block cell colour

    ws["A1"], ws["H3"], ws["H4"] = info["uwi"], "=(H10-E3)*I12+I10", "=(H10-E3)*J12+J10"
    ws["B3"], ws["E3"], ws["B5"], ws["E5"] = float(info["lati"]), -abs(float(info["long"])), info["kb"], info["gl"]
    ws["B7"] = info["mud_weight"][0]
    if info["mud_weight"][1]:
        ws["D7"] = "Assumed"

    df = info["formation_tops"]
    formation_tvd_array = np.sort((df[["TVD"]].to_numpy(dtype=float)), axis=None)  # for inputing missing DT values
    for i, formation in enumerate(["TRbaldnnl", "TRhalfway", "TRdoig", "TRmntny_U",
                                   "Trmontney (TRmntny_M)", "TRmntny_L", "PRbelloy"]):
        index_df = df[df["Formation tops"] == formation].index[0]
        ws["A" + str(11 + i)] = formation
        ws["B" + str(11 + i)] = float(df.iloc[index_df]["TVD"])
        ws["C" + str(11 + i)] = float(df.iloc[index_df]["Elev"])
    # get data from info, enumerated formation names and formation tops should equal

    ws2 = wb.create_sheet("logs")  # edit sheets sequentially
    ws2.column_dimensions["H"].width = 12.89 + width_const
    ws2.column_dimensions["I"].width = 12.89 + width_const
    ws2.column_dimensions["N"].width = 6.78 + width_const

    ws2["F1"], ws2["H1"], ws2["M1"], ws2["N1"] = "calc frm DT", "logs", "core test", "core test"
    ws2["A2"], ws2["B2"], ws2["C2"], ws2["D2"], ws2["E2"], ws2["F2"], ws2["G2"], ws2["H2"], \
        ws2["I2"], ws2["J2"], ws2["K2"], ws2["L2"], ws2["M2"], ws2["N2"], ws2["O2"], ws2["P2"] = \
        "DEPT", "GR", "RHOZ", "DTCO", "DTSM", "RHOZ", "DEPT", "RHOZ", \
        "EHD2_PPC1", "EHD1_PPC1", "P1AZ", "BS", "DEPT", "DEPT_shift", "den", "den "
    ws2["A3"], ws2["B3"], ws2["C3"], ws2["D3"], ws2["E3"], ws2["F3"], ws2["G3"], ws2["H3"], \
        ws2["I3"], ws2["J3"], ws2["K3"], ws2["L3"], ws2["M3"], ws2["N3"], ws2["O3"], ws2["P3"] = \
        "m", "gapi", "k/m^3", "us/m", "us/m", "k/m^3", "m", \
        "k/m^3", "mm", "mm", "deg", "mm", "m", "m", "g/cm^3", "kg/m^3"
    dept_col, dt_col, gr_col, rhoz_col = [las_df.columns.get_loc(col) if col in las_df else None
                                          for col in ["DEPT", "DT", "GR", "RHOZ"]]
    # get columns (may have been a good idea to check format in the first program)
    i = 4  # counter for row index
    dt_flag = False
    prev_dept = 0
    # workaround to circular reference error
    # instead of invalid value handling in Excel control cell direction fail-over in Python
    for r in dataframe_to_rows(las_df, index=False, header=False):
        if dept_col is not None:
            dept = r[dept_col]
        else:
            dept = None
        if gr_col is not None:
            gr = r[gr_col]
        else:
            gr = None
        if rhoz_col is not None:
            rhoz = r[rhoz_col]
        else:
            rhoz = None
        if dt_col is not None and not np.isnan(r[dt_col]):
            dt = r[dt_col]
            dt_flag = True
        else:
            dt = None
        dt_string, formation_tvd_array = missing_dt_fill(i, dept, prev_dept, formation_tvd_array)
        ws2.append({"A": dept,
                    "B": gr,
                    "C": rhoz or "=IFERROR(F" + str(i) + ", C" + (str(i - 1) if dt_flag else str(i + 1)) + ")",
                    # "=IFERROR(F" + str(i) + ",IF(C" + str(i + 1) + "=0, C" + str(i - 1) + ", C" + str(i + 1) + "))"
                    # Excel does not like circular references
                    "D": dt or (dt_string if dt_string == "NaN" else "=D" + (dt_string if dt_flag else str(i + 1))),
                    # shallow wells do not need checking
                    "E": "=" + str(coefficient_df["DTCO^2*DEPT"][0]) + "*(D" + str(i) + "^2*A" + str(i) + ")+" +
                         str(coefficient_df["DTCO^2"][0]) + "*(D" + str(i) + "^2)+" +
                         str(coefficient_df["DTCO"][0]) + "*(D" + str(i) + ")",
                    # "E": coefficient_df["DTCO^2*DEPT"][0] * (dt ** 2 * dept) + coefficient_df["DTCO^2"][0] * (dt **
                    # 2) + coefficient_df["DTCO"][0] * dt,
                    "F": "=(2.75-2.11*((D" + str(i) + "/3.28084-56)/(D" + str(i) + "/3.28084+200)))*1000*0.985",
                    # "F": (2.75-2.11*((dt/3.28084-56)/(dt/3.28084+200)))*1000*0.985,
                    # storing final value instead of formula takes more space
                    })
        i += 1
        prev_dept = dept
        # append log data to sheet

    x_values = Reference(ws2, min_col=1, min_row=4, max_row=i)

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.x_axis.scaling.min = 1000
    chart.x_axis.scaling.max = 3000
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws2, min_col=3, min_row=4, max_row=i)
    series = Series(x_values, values, title="RHOZ")  # both logging and DT calc?
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "ed7d31"  # orange
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    values = Reference(ws2, min_col=6, min_row=4, max_row=i)
    series = Series(x_values, values, title="RHOZ_cal_frm_DT")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "4472c4"  # blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.height = 17
    chart.width = 9.5
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.3, y=0.8, h=0.1, w=0.4))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws2.add_chart(chart, "G4")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 200
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws2, min_col=2, min_row=4, max_row=i)
    series = Series(x_values, values, title="GR")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="4472c4"))  # blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "4472c4"  # blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    chart.series.append(series)
    chart.height = 17
    chart.width = 9.5
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.3, y=0.8, h=0.1, w=0.4))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws2.add_chart(chart, "L4")

    ws3 = wb.create_sheet("mechanics")

    ws3.column_dimensions["A"].width = 11.33 + width_const
    ws3.column_dimensions["B"].width = 12.33 + width_const
    ws3.column_dimensions["F"].width = 11.11 + width_const
    ws3.column_dimensions["G"].width = 6.56 + width_const
    ws3.column_dimensions["H"].width = 7.22 + width_const
    ws3.column_dimensions["J"].width = 11.22 + width_const
    ws3.column_dimensions["K"].width = 12.33 + width_const
    ws3.column_dimensions["L"].width = 10.22 + width_const
    ws3.column_dimensions["M"].width = 9.67 + width_const
    ws3.freeze_panes = "A14"

    ws3["A2"], ws3["B2"], ws3["C2"], ws3["A3"], ws3["B3"], ws3["C3"], ws3["A4"], ws3["B4"], ws3["C4"] = \
        "EKB", "EGL", "assumed shallow density", "m", "m", "k/m^3", "='basic infor'!B5", "='basic infor'!E5", \
        "=INDEX(logs!$C$4:logs!$C$" + str(i - 1) + ",MATCH(TRUE,INDEX((logs!$C$4:logs!$C$" + str(i - 1) + "<>0),0),0))"
    ws3["A6"], ws3["B6"], ws3["A7"], ws3["B7"], ws3["A8"] = \
        "# of records", "Max depth (m)", "=COUNT(logs!A4:A" + str(i - 1) + ")", "=INDEX($A$14:$A$" + str(
            i + 9) + ",A7)", "=A7+14"  # remember that Excel indexing starts at 1
    ws3["J1"], ws3["J2"], ws3["J3"], ws3["J4"], ws3["J5"], ws3["J6"], ws3["J7"], ws3["J8"], ws3["J9"], ws3["J10"] = \
        "Tops", "surface", "TRbaldnnl", "TRhalfway", "TRdoig", "TRmntny_U", "TRmntny_M", "TRmntny_L", "PRbelloy", "TD"
    ws3["K1"], ws3["K3"], ws3["K4"], ws3["K5"], ws3["K6"], ws3["K7"], ws3["K8"], ws3["K9"], ws3["K10"] = \
        "DEPT", "='basic infor'!B11", "='basic infor'!B12", "='basic infor'!B13", \
        "='basic infor'!B14", "='basic infor'!B15", "='basic infor'!B16", "='basic infor'!B17", "=B7"
    ws3["L1"], ws3["L3"], ws3["L4"], ws3["L5"], ws3["L6"], ws3["L7"], ws3["L8"], ws3["L9"], ws3["L10"] = \
        "record #", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K3)), ABS($A$14:$A$" + str(i + 9) + "-K3),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K4)), ABS($A$14:$A$" + str(i + 9) + "-K4),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K5)), ABS($A$14:$A$" + str(i + 9) + "-K5),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K6)), ABS($A$14:$A$" + str(i + 9) + "-K6),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K7)), ABS($A$14:$A$" + str(i + 9) + "-K7),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K8)), ABS($A$14:$A$" + str(i + 9) + "-K8),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K9)), ABS($A$14:$A$" + str(i + 9) + "-K9),0)+13", \
        "=MATCH(MIN(ABS($A$14:$A$" + str(i + 9) + "-K10)), ABS($A$14:$A$" + str(i + 9) + "-K10),0)+13"
    for record in range(3, 11):  # inputting array formulas prior to openpyxl version 3.1.0
        ws3.formula_attributes["L" + str(record)] = {"t": "array", "ref": "L" + str(record)}
    for row_count, row in enumerate(ws3["N1":"R9"]):
        for col_count, cell in enumerate(row):
            cell.value = ["Esta coeff", "nu coef", "UCS coef", "TSTR coef", "FA coef",
                          0.75, 1, 0.45, 0.06, 0.6,
                          0.75, 1, 0.45, 0.06, 0.6,
                          1.6, 1, 0.6, 0.12, 0.82,
                          1.6, 1, 0.6, 0.12, 0.82,
                          1.6, 1, 0.6, 0.12, 0.82,
                          1.6, 1, 0.6, 0.12, 0.82,
                          1.6, 1, 0.6, 0.12, 0.82,
                          0.75, 1, 0.45, 0.06, 0.6][row_count * 5 + col_count]

    for row_count, row in enumerate(ws3["A12":"R13"]):
        for col_count, cell in enumerate(row):
            cell.value = ["DEPT", "GR", "RHOZ", "DTCO", "DTSM", "DTCO", "DTSM", "Sig_V", "Sig_V_G", "Gdyn", "Kdyn",
                          "Edyn", "nu_dyn", "Esta", "nu_sta", "UCS", "TSTR", "FA",
                          "m", "gapi", "k/m^3", "us/m", "us/m", "us/ft", "us/ft", "MPa", "kPa/m", "MPsi", "MPsi", "GPa",
                          None, "GPa", None, "MPa", "MPa", "Deg"][row_count * 18 + col_count]
    # ws3["T10"], ws3["T12"],  ws3["T13"], ws3["U12"],  ws3["U13"], ws3["V12"], ws3["W12"],  ws3["W13"], ws3["X12"],  \
    #     ws3["X13"] = "core test", "DEPT", "m", "E", "GPa", "nv", "UCS", "MPa", "FA", "deg"
    # core test unnecessary

    for i_ws3 in range(4, i):
        ws3.append(
            {"A": "=IF(AND(ROW(A" + str(i_ws3 + 10) + ")<$A$8,logs!A" + str(i_ws3) + ">0), logs!A" + str(i_ws3) + ")",
             "B": "=IF(AND(ROW(B" + str(i_ws3 + 10) + ")<$A$8,logs!B" + str(i_ws3) + ">0), logs!B" + str(i_ws3) + ")",
             "C": "=IF(AND(ROW(C" + str(i_ws3 + 10) + ")<$A$8,logs!C" + str(i_ws3) + ">0), logs!C" + str(i_ws3) + ")",
             "D": "=IF(AND(ROW(D" + str(i_ws3 + 10) + ")<$A$8,logs!D" + str(i_ws3) + ">0), logs!D" + str(i_ws3) + ")",
             "E": "=IF(AND(ROW(E" + str(i_ws3 + 10) + ")<$A$8,logs!E" + str(i_ws3) + ">0), logs!E" + str(i_ws3) + ")",
             "F": "=D" + str(i_ws3 + 10) + "/3.28084",
             "G": "=E" + str(i_ws3 + 10) + "/3.28084",
             "H": "=$C$4*($A$14-($A$4-$B$4))*9.81/1000000" if i_ws3 == 4 else
             "=H" + str(i_ws3 + 9) + "+(A" + str(i_ws3 + 10) + "-A" + str(i_ws3 + 9) + ")*9.81*C" + str(
                 i_ws3 + 10) + "/1000000",
             "I": "=H" + str(i_ws3 + 10) + "/A" + str(i_ws3 + 10) + "*1000",
             "J": "=13474.45*C" + str(i_ws3 + 10) + "/1000/G" + str(i_ws3 + 10) + "^2",
             "K": "=13474.45*C" + str(i_ws3 + 10) + "/1000/F" + str(i_ws3 + 10) + "^2-4/3*J" + str(i_ws3 + 10),
             "L": "=(9*J" + str(i_ws3 + 10) + "*K" + str(i_ws3 + 10) + ")/(J" + str(i_ws3 + 10) + "+K" + str(i_ws3 + 10)
                  + "*3)*6.89475729",
             "M": "=(3*K" + str(i_ws3 + 10) + "-2*J" + str(i_ws3 + 10) + ")/(6*K" + str(i_ws3 + 10) + "+2*J"
                  + str(i_ws3 + 10) + ")",
             "N": "=(0.032*L" + str(i_ws3 + 10) + "^1.632)*IF(ROW(L" + str(i_ws3 + 10) + ")<$L$3,$N$2,IF(AND(ROW(L" +
                  str(i_ws3 + 10) + ")>=$L$3,ROW(L" + str(i_ws3 + 10) + ")<$L$4),$N$3,IF(AND(ROW(L" +
                  str(i_ws3 + 10) + ")>=$L$4,ROW(L" + str(i_ws3 + 10) + ")<$L$5),$N$4,IF(AND(ROW(L" +
                  str(i_ws3 + 10) + ")>=$L$5,ROW(L" + str(i_ws3 + 10) + ")<$L$6),$N$5,IF(AND(ROW(L" +
                  str(i_ws3 + 10) + ")>=$L$6,ROW(L" + str(i_ws3 + 10) + ")<$L$7),$N$6,IF(AND(ROW(L" +
                  str(i_ws3 + 10) + ")>=$L$7,ROW(L" + str(i_ws3 + 10) + ")<$L$8),$N$7,IF(AND(ROW(L" +
                  str(i_ws3 + 10) + ")>=$L$8,ROW(L" + str(i_ws3 + 10) + ")<$L$9),$N$8,$N$9)))))))",
             "O": "=M" + str(i_ws3 + 10) + "*IF(ROW(M" + str(i_ws3 + 10) + ")<$L$3,$O$2,IF(AND(ROW(M" +
                  str(i_ws3 + 10) + ")>=$L$3,ROW(M" + str(i_ws3 + 10) + ")<$L$4),$O$3,IF(AND(ROW(M" +
                  str(i_ws3 + 10) + ")>=$L$4,ROW(M" + str(i_ws3 + 10) + ")<$L$5),$O$4,IF(AND(ROW(M" +
                  str(i_ws3 + 10) + ")>=$L$5,ROW(M" + str(i_ws3 + 10) + ")<$L$6),$O$5,IF(AND(ROW(M" +
                  str(i_ws3 + 10) + ")>=$L$6,ROW(M" + str(i_ws3 + 10) + ")<$L$7),$O$6,IF(AND(ROW(M" +
                  str(i_ws3 + 10) + ")>=$L$7,ROW(M" + str(i_ws3 + 10) + ")<$L$8),$O$7,IF(AND(ROW(M" +
                  str(i_ws3 + 10) + ")>=$L$8,ROW(M" + str(i_ws3 + 10) + ")<$L$9),$O$8,$O$9)))))))",
             "P": "=(N" + str(i_ws3 + 10) + "*4.242)*IF(ROW(N" + str(i_ws3 + 10) + ")<$L$3,$P$2,IF(AND(ROW(N" +
                  str(i_ws3 + 10) + ")>=$L$3,ROW(N" + str(i_ws3 + 10) + ")<$L$4),$P$3,IF(AND(ROW(N" +
                  str(i_ws3 + 10) + ")>=$L$4,ROW(N" + str(i_ws3 + 10) + ")<$L$5),$P$4,IF(AND(ROW(N" +
                  str(i_ws3 + 10) + ")>=$L$5,ROW(N" + str(i_ws3 + 10) + ")<$L$6),$P$5,IF(AND(ROW(N" +
                  str(i_ws3 + 10) + ")>=$L$6,ROW(N" + str(i_ws3 + 10) + ")<$L$7),$P$6,IF(AND(ROW(N" +
                  str(i_ws3 + 10) + ")>=$L$7,ROW(N" + str(i_ws3 + 10) + ")<$L$8),$P$7,IF(AND(ROW(N" +
                  str(i_ws3 + 10) + ")>=$L$8,ROW(N" + str(i_ws3 + 10) + ")<$L$9),$P$8,$P$9)))))))",
             "Q": "=P" + str(i_ws3 + 10) + "*IF(ROW(P" + str(i_ws3 + 10) + ")<$L$3,$Q$2,IF(AND(ROW(P" +
                  str(i_ws3 + 10) + ")>=$L$3,ROW(P" + str(i_ws3 + 10) + ")<$L$4),$Q$3,IF(AND(ROW(P" +
                  str(i_ws3 + 10) + ")>=$L$4,ROW(P" + str(i_ws3 + 10) + ")<$L$5),$Q$4,IF(AND(ROW(P" +
                  str(i_ws3 + 10) + ")>=$L$5,ROW(P" + str(i_ws3 + 10) + ")<$L$6),$Q$5,IF(AND(ROW(P" +
                  str(i_ws3 + 10) + ")>=$L$6,ROW(P" + str(i_ws3 + 10) + ")<$L$7),$Q$6, IF(AND(ROW(P" +
                  str(i_ws3 + 10) + ")>=$L$7,ROW(P" + str(i_ws3 + 10) + ")<$L$8),$Q$7, IF(AND(ROW(P" +
                  str(i_ws3 + 10) + ")>=$L$8,ROW(P" + str(i_ws3 + 10) + ")<$L$9), $Q$8, $Q$9)))))))",
             "R": "=ASIN(((1/D" + str(i_ws3 + 10) + "*1000000)-1000)/((1/D" + str(i_ws3 + 10) +
                  "*1000000)+1000))/PI()*180*IF(ROW(D" + str(i_ws3 + 10) + ")<$L$3,$R$2,IF(AND(ROW(D" +
                  str(i_ws3 + 10) + ")>=$L$3,ROW(D" + str(i_ws3 + 10) + ")<$L$4),$R$3,IF(AND(ROW(D" +
                  str(i_ws3 + 10) + ")>=$L$4,ROW(D" + str(i_ws3 + 10) + ")<$L$5),$R$4,IF(AND(ROW(D" +
                  str(i_ws3 + 10) + ")>=$L$5,ROW(D" + str(i_ws3 + 10) + ")<$L$6),$R$5,IF(AND(ROW(D" +
                  str(i_ws3 + 10) + ")>=$L$6,ROW(D" + str(i_ws3 + 10) + ")<$L$7),$R$6,IF(AND(ROW(D" +
                  str(i_ws3 + 10) + ")>=$L$7,ROW(D" + str(i_ws3 + 10) + ")<$L$8),$R$7, IF(AND(ROW(D" +
                  str(i_ws3 + 10) + ")>=$L$8,ROW(D" + str(i_ws3 + 10) + ")<$L$9), $R$8,$R$9)))))))"})

    for row in range(0, ws3.max_row + 1):
        ws3.row_dimensions[row].height = 15  # adjust height of rows after all have been generated

    x_values = Reference(ws3, min_col=1, min_row=14, max_row=i+10)

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "Young's Modulus (GPa)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = float(las_df["DEPT"].iloc[0]) // 100 * 100
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 100
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws3, min_col=14, min_row=14, max_row=i+10)
    series = Series(x_values, values, title="E")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="c8500a"))  # brown
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 14.5
    chart.width = 9.6
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.6, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws3.add_chart(chart, "A16")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "Poisson's ratio"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = float(las_df["DEPT"].iloc[0]) // 100 * 100
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 0.5
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws3, min_col=15, min_row=14, max_row=i + 10)
    series = Series(x_values, values, title="PR")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="2e75b6"))  # lighter blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 14.5
    chart.width = 9.6
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.6, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws3.add_chart(chart, "F16")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "UCS (MPa)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = float(las_df["DEPT"].iloc[0]) // 100 * 100
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 300
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws3, min_col=16, min_row=14, max_row=i + 10)
    series = Series(x_values, values, title="UCS_calc")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="7030a0"))  # purple
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 1.25 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 14.5
    chart.width = 9.6
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.6, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws3.add_chart(chart, "K16")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "FA (deg)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = float(las_df["DEPT"].iloc[0]) // 100 * 100
    chart.x_axis.scaling.min = 0  # no max
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws3, min_col=18, min_row=14, max_row=i + 10)
    series = Series(x_values, values, title="FA_calc")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="70ad47"))  # light green
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 1.25 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 14.5
    chart.width = 9.6
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.6, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws3.add_chart(chart, "P16")

    ws4 = wb.create_sheet("Pore pressure Bower")

    ws4.sheet_view.zoomScale = 84

    ws4.column_dimensions["D"].width = 10.78 + width_const
    ws4.column_dimensions["E"].width = 9.67 + width_const
    ws4.column_dimensions["F"].width = 16.78 + width_const
    ws4.column_dimensions["G"].width = 16.89 + width_const
    ws4.column_dimensions["H"].width = 14.78 + width_const
    ws4.column_dimensions["I"].width = 9.67 + width_const
    ws4.column_dimensions["J"].width = 9.67 + width_const
    ws4.column_dimensions["K"].width = 9.67 + width_const
    ws4.column_dimensions["L"].width = 9.67 + width_const
    ws4.column_dimensions["O"].width = 13.89 + width_const
    ws4.column_dimensions["P"].width = 12.78 + width_const
    ws4.column_dimensions["Q"].width = 14.56 + width_const
    ws4.column_dimensions["R"].width = 13.33 + width_const
    ws4.column_dimensions["S"].width = 12.89 + width_const
    ws4.freeze_panes = "A13"

    ws4["A1"], ws4["B1"], ws4["D1"], ws4["E1"], ws4["A2"], ws4["B2"], ws4["D2"], ws4["E2"] = \
        "DT_shale", "P_v_shale", "Vmax", "SigV_eff_max", "us/ft", "ft/s", "ft/s", "psi"
    ws4["A3"], ws4["D3"], ws4["E3"], ws4["A4"], ws4["B4"], ws4["G1"], ws4["H1"], ws4["I1"] = \
        "61-91", 18000, "=(($D$3-5000)/$G$2)^(1/$H$2)", 76, "=1/A4*1000*1000", "A", "B", "U"
    ws4["G2"], ws4["H2"], ws4["I2"] = 16, 0.808, 1.1
    ws4["K1"], ws4["K3"], ws4["K4"], ws4["K5"], ws4["K6"], ws4["K7"], ws4["K8"], ws4["K9"] = \
        "Tops", "TRhalfway", "TRdoig", "TRmntny_U", "TRmntny_M", "TRmntny_L", "PRbelloy", "TD"
    ws4["L1"], ws4["L2"], ws4["L3"], ws4["L4"], ws4["L5"], ws4["L6"], ws4["L7"], ws4["L8"], ws4["L9"] = \
        "DEPT", "m", "=mechanics!K4", "=mechanics!K5", \
        "=mechanics!K6", "=mechanics!K7", "=mechanics!K8", "=mechanics!K9", "=mechanics!K10"
    ws4["M1"], ws4["M3"], ws4["M4"], ws4["M5"], ws4["M6"], ws4["M7"], ws4["M8"], ws4["M9"] = \
        "record #", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L3)), ABS($A$13:$A$" + str(i + 9) + "-L3),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L4)), ABS($A$13:$A$" + str(i + 9) + "-L4),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L5)), ABS($A$13:$A$" + str(i + 9) + "-L5),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L6)), ABS($A$13:$A$" + str(i + 9) + "-L6),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L7)), ABS($A$13:$A$" + str(i + 9) + "-L7),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L8)), ABS($A$13:$A$" + str(i + 9) + "-L8),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-L9)), ABS($A$13:$A$" + str(i + 9) + "-L9),0)+12",
    for record in range(3, 10):  # inputting array formulas prior to openpyxl version 3.1.0
        ws4.formula_attributes["M" + str(record)] = {"t": "array", "ref": "M" + str(record)}
    ws4["N1"], ws4["N2"], ws4["N3"], ws4["N4"], ws4["N5"], ws4["N6"], ws4["N7"], ws4["N8"] = "Pp_G", "kPa/m", \
        "=AVERAGE(INDEX($L$13:$L$" + str(i + 9) + ",M3-12):INDEX($L$13:$L$" + str(i + 9) + ",M4-12))", \
        "=AVERAGE(INDEX($L$13:$L$" + str(i + 9) + ",M4-12):INDEX($L$13:$L$" + str(i + 9) + ",M5-12))", \
        "=AVERAGE(INDEX($L$13:$L$" + str(i + 9) + ",M5-12):INDEX($L$13:$L$" + str(i + 9) + ",M6-12))", \
        "=AVERAGE(INDEX($L$13:$L$" + str(i + 9) + ",M6-12):INDEX($L$13:$L$" + str(i + 9) + ",M7-12))", \
        "=AVERAGE(INDEX($L$13:$L$" + str(i + 9) + ",M7-12):INDEX($L$13:$L$" + str(i + 9) + ",M8-12))", \
        "=AVERAGE(INDEX($L$13:$L$" + str(i + 9) + ",M8-12):INDEX($L$13:$L$" + str(i + 9) + ",M9-12))",

    for row_count, row in enumerate(ws4["A11":"L12"]):
        for col_count, cell in enumerate(row):
            cell.value = ["DEPT", "Sig_V", "DTCO", "P_wave_v", "Pp vergin", "Sig_V_eff_vergin", "P_wave_curve",
                          "SigV_eff_curve", "SigV_eff", "Pp_bower", "Pp", "Pp_G",
                          "m", "MPa", "us/ft", "ft/s", "MPa", "psi", "ft/s", "psi", "psi", "psi",
                          "MPa", "kPa/m"][row_count * 12 + col_count]

    halfway_dept = float(df.iloc[df[df["Formation tops"] == "TRhalfway"].index[0]]["TVD"])
    halfway_len = len(las_df[las_df["DEPT"] < halfway_dept])
    for i_ws4 in range(4, i - halfway_len):  # do not overcount
        p_wave_curve = "=5000+$G$2*H" + str(i_ws4 + 9) + "^$H$2" if (i_ws4 - 4) * 200 <= 20000 else None
        sig_v_eff_curve = (i_ws4 - 4) * 200 if (i_ws4 - 4) * 200 <= 20000 else None
        ws4.append(
            {"A": "=INDEX(mechanics!$A$14:$A$" + str(i + 10) + ",ROW(mechanics!A" + str(i_ws4 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "B": "=INDEX(mechanics!$H$14:$H$" + str(i + 10) + ",ROW(mechanics!H" + str(i_ws4 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "C": "=INDEX(mechanics!$F$14:$F$" + str(i + 10) + ",ROW(mechanics!F" + str(i_ws4 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "D": "=1/C" + str(i_ws4 + 9) + "*1000*1000",
             "E": "=A" + str(i_ws4 + 9) + "*1000*9.81/1000/1000",
             "F": "=(B" + str(i_ws4 + 9) + "-E" + str(i_ws4 + 9) + ")*145.038",
             "G": p_wave_curve,
             "H": sig_v_eff_curve,
             "I": "=((((D" + str(i_ws4 + 9) + "-5000)/$G$2)^(1/$H$2))/$E$3)^$I$2*$E$3",
             "J": "=B" + str(i_ws4 + 9) + "*145.038-I" + str(i_ws4 + 9),
             "K": "=J" + str(i_ws4 + 9) + "/145.038",
             "L": "=K" + str(i_ws4 + 9) + "/A" + str(i_ws4 + 9) + "*1000"
             })

    for row_count, row in enumerate(ws4["G1":"H113"]):  # can probably assume there are enough measurements in range
        for col_count, cell in enumerate(row):
            cell.fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")

    for row in range(0, ws4.max_row + 1):
        ws4.row_dimensions[row].height = 15  # adjust height of rows after all have been generated

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = halfway_dept // 100 * 100
    chart.x_axis.title = "P_wave_v (ft/s)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.x_axis.scaling.min = 10000  # no max
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws4, min_col=4, min_row=13, max_row=i-halfway_len+8)
    x_values = Reference(ws4, min_col=1, min_row=13, max_row=i-halfway_len+8)
    series = Series(x_values, values, title="P_wave_v")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "4472c4"  # blue
    series.marker.graphicalProperties.line.solidFill = "00b0f0"  # light blue
    series.marker.size = 2
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.height = 15.5
    chart.width = 11.6
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws4.add_chart(chart, "B15")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = halfway_dept // 100 * 100
    chart.x_axis.title = "Pressure gradient (kPa/m)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.x_axis.scaling.min = 0  # no max
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws4, min_col=12, min_row=13, max_row=i-halfway_len+8)
    x_values = Reference(ws4, min_col=1, min_row=13, max_row=i-halfway_len+8)
    series = Series(x_values, values, title="Pore pressure gradient Bower")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "4472c4"  # blue
    series.marker.graphicalProperties.line.solidFill = "00b0f0"  # light blue
    series.marker.size = 2
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.height = 18.7
    chart.width = 16.5
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws4.add_chart(chart, "I15")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "Effective stress (psi)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.title = "P-wave velocity (ft/s)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 4000
    chart.y_axis.scaling.min = 5000
    chart.y_axis.scaling.max = 19000
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws4, min_col=7, min_row=13, max_row=i - halfway_len + 8)
    x_values = Reference(ws4, min_col=8, min_row=13, max_row=i - halfway_len + 8)
    series = Series(values, x_values, title="normal compaction curve")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="7f7f7f"))  # gray
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.25 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws4, min_col=4, min_row=13, max_row=i - halfway_len + 8)
    x_values = Reference(ws4, min_col=9, min_row=13, max_row=i - halfway_len + 8)
    series = Series(values, x_values, title="P_wave_v")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "ffc000"  # blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.height = 4.5
    chart.width = 10
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.05, h=0.8, w=0.75))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.5, y=0.5, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws4.add_chart(chart, "F2")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = halfway_dept // 100 * 100
    chart.x_axis.title = "Pressure (MPa)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.x_axis.scaling.min = 0  # no max
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws4, min_col=11, min_row=13, max_row=i - halfway_len + 8)
    x_values = Reference(ws4, min_col=1, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Pore pressure gradient Bower")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "4472c4"  # blue
    series.marker.graphicalProperties.line.solidFill = "00b0f0"  # light blue
    series.marker.size = 2
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.height = 18.7
    chart.width = 16.5
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.8, h=0.1, w=0.3))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws4.add_chart(chart, "Q15")

    ws5 = wb.create_sheet("stress")

    ws5.sheet_view.zoomScale = 85

    ws5.column_dimensions["C"].width = 11.89 + width_const
    ws5.column_dimensions["D"].width = 11.33 + width_const
    ws5.column_dimensions["F"].width = 15.22 + width_const
    ws5.column_dimensions["K"].width = 6.22 + width_const
    ws5.column_dimensions["L"].width = 9.67 + width_const
    ws5.column_dimensions["N"].width = 11.22 + width_const
    ws5.column_dimensions["P"].width = 10.33 + width_const
    ws5.column_dimensions["Q"].width = 10.33 + width_const
    ws5.column_dimensions["R"].width = 11.22 + width_const
    ws5.column_dimensions["S"].width = 9.67 + width_const
    ws5.column_dimensions["T"].width = 11.33 + width_const
    ws5.freeze_panes = "A13"

    ws5["A1"], ws5["B1"], ws5["C1"], ws5["A3"], ws5["A4"], ws5["B4"], ws5["A5"], ws5["B5"] = \
        "mud weight", "='basic infor'!B7", "kg/m^3", \
        "Strain", "strain x", "='basic infor'!H3*$B$8", "strain y", "='basic infor'!H4*$B$9"
    ws5["A7"], ws5["A8"], ws5["B8"], ws5["A9"], ws5["B9"] = "multiply", "x", 0.85, "y", 0.85
    ws5["F1"], ws5["F2"], ws5["F3"], ws5["F4"], ws5["F5"], ws5["F6"], ws5["F7"], ws5["F8"], ws5["F9"] = \
        "Pp interval", "above", "TRhalfway", "TRdoig", "TRmntny_U", "Trmontney (m for BC wells)", "TRmntny_L", \
        "PRbelloy", "TD"
    ws5["G1"], ws5["H1"], ws5["I1"], ws5["J1"] = "MD (m)", "record #", "gradient (kPa)", "gradient from Bower (kPa)"
    ws5["G3"], ws5["G4"], ws5["G5"], ws5["G6"], ws5["G7"], ws5["G8"], ws5["G9"] = \
        "='basic infor'!B12", "='basic infor'!B13", "='basic infor'!B14", "='basic infor'!B15", \
        "='Pore pressure Bower'!L7", "='Pore pressure Bower'!L8", "='Pore pressure Bower'!L9"
    ws5["H3"], ws5["H4"], ws5["H5"], ws5["H6"], ws5["H7"], ws5["H8"], ws5["H9"] = \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G3)), ABS($A$13:$A$" + str(i + 9) + "-G3),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G4)), ABS($A$13:$A$" + str(i + 9) + "-G4),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G5)), ABS($A$13:$A$" + str(i + 9) + "-G5),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G6)), ABS($A$13:$A$" + str(i + 9) + "-G6),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G7)), ABS($A$13:$A$" + str(i + 9) + "-G7),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G8)), ABS($A$13:$A$" + str(i + 9) + "-G8),0)+12", \
        "=MATCH(MIN(ABS($A$13:$A$" + str(i + 9) + "-G9)), ABS($A$13:$A$" + str(i + 9) + "-G9),0)+12",
    for record in range(3, 10):  # inputting array formulas prior to openpyxl version 3.1.0
        ws5.formula_attributes["H" + str(record)] = {"t": "array", "ref": "H" + str(record)}
    ws5["I3"], ws5["I4"], ws5["I5"], ws5["I6"], ws5["I7"], ws5["I8"] = "=J3", "=J4", "=J5", "=J6", "=J7", "=J8"
    ws5["J3"], ws5["J4"], ws5["J5"], ws5["J6"], ws5["J7"], ws5["J8"] = \
        "='Pore pressure Bower'!N3", "='Pore pressure Bower'!N4", "='Pore pressure Bower'!N5", \
        "='Pore pressure Bower'!N6", "='Pore pressure Bower'!N7", "='Pore pressure Bower'!N8"

    ws5["O1"] = "Stress gradients"
    for row_count, row in enumerate(ws5["N2":"Q6"]):
        for col_count, cell in enumerate(row):
            cell.value = [None, "Sig_v_G", "Sig_h_G", "sig_H_G",
                          "All",
                          "=AVERAGE($O$13:$O$" + str(i + 9) + ")",
                          "=AVERAGE($P$13:$P$" + str(i + 9) + ")",
                          "=AVERAGE($Q$13:$Q$" + str(i + 9) + ")",
                          "Montney",
                          "=AVERAGE(INDEX($O$13:$O$" + str(i + 9) + ",$H$5-12):INDEX($O$13:$O$" + str(i + 9) + ",$H$8"
                                                                                                               "-12))",
                          "=AVERAGE(INDEX($P$13:$P$" + str(i + 9) + ",$H$5-12):INDEX($P$13:$P$" + str(i + 9) + ",$H$8"
                                                                                                               "-12))",
                          "=AVERAGE(INDEX($Q$13:$Q$" + str(i + 9) + ",$H$5):INDEX($Q$13:$Q$" + str(i + 9) + ",$H$8))",
                          "TRmntny_U",
                          "=AVERAGE(INDEX($O$13:$O$" + str(i + 9) + ",$H$5-12):INDEX($O$13:$O$" + str(i + 9) + ",$H$6"
                                                                                                               "-12))",
                          "=AVERAGE(INDEX($P$13:$P$" + str(i + 9) + ",$H$5-12):INDEX($P$13:$P$" + str(i + 9) + ",$H$6"
                                                                                                               "-12))",
                          "=AVERAGE(INDEX($Q$13:$Q$" + str(i + 9) + ",$H$5):INDEX($Q$13:$Q$" + str(i + 9) + ",$H$6))",
                          "TRmntny_M and Trmntny_L",
                          "=AVERAGE(INDEX($O$13:$O$" + str(i + 9) + ",$H$6-12):INDEX($O$13:$O$" + str(i + 9) + ",$H$8"
                                                                                                               "-12))",
                          "=AVERAGE(INDEX($P$13:$P$" + str(i + 9) + ",$H$6-12):INDEX($P$13:$P$" + str(i + 9) + ",$H$8"
                                                                                                               "-12))",
                          "=AVERAGE(INDEX($Q$13:$Q$" + str(i + 9) + ",$H$6-12):INDEX($Q$13:$Q$" + str(i + 9) + ",$H$8"
                                                                                                               "-12))"][
                row_count * 4 + col_count]

    for row_count, row in enumerate(ws5["A11":"U12"]):
        for col_count, cell in enumerate(row):
            cell.value = ["DEPT", "Sig_V", "Pp", "Esta", "nu_sta", "UCS", "TSTR", "FA", "Sig_h", "Sig_H",
                          "ECD pressure", "Breakout", "induce frac", "Pp_G", "Sig_v_G", "Sig_h_G", "sig_H_G",
                          "DFIT_dept", "DFIT_value", "Pp_test_depth", "PP_test",
                          "m", "MPa", "MPa", "GPa", None, "MPa", "MPa", "DEG", "MPa", "MPa", "MPa", None, None, "kPa/m",
                          "kPa/m", "kPa/m", "kPa/m", "m", "MPa", "m", "MPa"][row_count * 21 + col_count]

    for i_ws5 in range(4, i - halfway_len):  # do not overcount
        ws5.append(
            {"A": "=INDEX(mechanics!$A$14:$A$" + str(i + 10) + ",ROW(mechanics!A" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "B": "=INDEX(mechanics!$H$14:$H$" + str(i + 10) + ",ROW(mechanics!H" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "C": "=A13/1000*IF(ROW(A" + str(i_ws5 + 9) + ")<$H$3,$I$2,IF(AND(ROW(A" + str(i_ws5 + 9) + ")>=$H$3,ROW(A"
                  + str(i_ws5 + 9) + ")<$H$4),$I$3,"
                                     "IF(AND(ROW(A" + str(i_ws5 + 9) + ")>=$H$4,ROW(A" + str(
                 i_ws5 + 9) + ")<$H$5),$I$4,"
                              "IF(AND(ROW(A" + str(i_ws5 + 9) + ")>=$H$5,ROW(A" + str(i_ws5 + 9) + ")<$H$6),$I$5,"
                                                                                                   "IF(AND(ROW(A" + str(
                 i_ws5 + 9) + ")>=$H$6,ROW(A" + str(i_ws5 + 9) + ")<$H$7),$I$6,"
                                                                 "IF(AND(ROW(A" + str(
                 i_ws5 + 9) + ")>=$H$7,ROW(A" + str(i_ws5 + 9) + ")<$H$8),$I$7,$I$8))))))",
             "D": "=INDEX(mechanics!$N$14:$N$" + str(i + 10) + ",ROW(mechanics!N" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "E": "=INDEX(mechanics!$O$14:$O$" + str(i + 10) + ",ROW(mechanics!O" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "F": "=INDEX(mechanics!$P$14:$P$" + str(i + 10) + ",ROW(mechanics!P" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "G": "=INDEX(mechanics!$Q$14:$Q$" + str(i + 10) + ",ROW(mechanics!Q" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "H": "=INDEX(mechanics!$R$14:$R$" + str(i + 10) + ",ROW(mechanics!R" + str(i_ws5 + 10) +
                  ")-14+mechanics!$L$4-13)",
             "I": "=E" + str(i_ws5 + 9) + "/(1-E" + str(i_ws5 + 9) + ")*B" + str(i_ws5 + 9) + "-E" + str(i_ws5 + 9) +
                  "/(1-E" + str(i_ws5 + 9) + ")*C" + str(i_ws5 + 9) + "+C" + str(i_ws5 + 9) + "+D" + str(i_ws5 + 9) +
                  "/(1-E" + str(i_ws5 + 9) + "^2)*$B$4+E" + str(i_ws5 + 9) + "*D" + str(i_ws5 + 9) + "/(1-E" +
                  str(i_ws5 + 9) + "^2)*$B$5",
             "J": "=E" + str(i_ws5 + 9) + "/(1-E" + str(i_ws5 + 9) + ")*B" + str(i_ws5 + 9) + "-E" + str(i_ws5 + 9) +
                  "/(1-E" + str(i_ws5 + 9) + ")*C" + str(i_ws5 + 9) + "+C" + str(i_ws5 + 9) + "+E" + str(i_ws5 + 9) +
                  "*D" + str(i_ws5 + 9) + "/(1-E" + str(i_ws5 + 9) + "^2)*$B$4+D" + str(i_ws5 + 9) + "/(1-E" +
                  str(i_ws5 + 9) + "^2)*$B$5",
             "K": "=$B$1*A" + str(i_ws5 + 9) + "*9.81/1000000*1.03",
             "L": "=3*J" + str(i_ws5 + 9) + "-I" + str(i_ws5 + 9) + "-K" + str(i_ws5 + 9) + "-C" + str(i_ws5 + 9) +
                  "-F" + str(i_ws5 + 9) + "-(TAN(PI()/4+(H" + str(i_ws5 + 9) + "*PI()/180)/2))^2*(K" + str(i_ws5 + 9) +
                  "-C" + str(i_ws5 + 9) + ")",
             "M": "=K" + str(i_ws5 + 9) + "-3*I" + str(i_ws5 + 9) + "+J" + str(i_ws5 + 9) + "-G" + str(i_ws5 + 9) +
                  "+C" + str(i_ws5 + 9) + "",
             "N": "=C" + str(i_ws5 + 9) + "/A" + str(i_ws5 + 9) + "*1000",
             "O": "=B" + str(i_ws5 + 9) + "/A" + str(i_ws5 + 9) + "*1000",
             "P": "=I" + str(i_ws5 + 9) + "/A" + str(i_ws5 + 9) + "*1000",
             "Q": "=J" + str(i_ws5 + 9) + "/A" + str(i_ws5 + 9) + "*1000"
             })

    for row in range(0, ws5.max_row + 1):
        ws5.row_dimensions[row].height = 15  # adjust height of rows after all have been generated

    ws5["I2"] = 10.1  # sensible default, alternatively 9.81 for hydrostatic

    x_values = Reference(ws5, min_col=1, min_row=13, max_row=i - halfway_len + 8)

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = halfway_dept // 100 * 100
    chart.x_axis.title = "Stress (MPa)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 80
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws5, min_col=3, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Pore pressure")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "00b0f0"  # light blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="00b0f0"))  # light blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=9, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Sig_h")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "00b050"  # dark green
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="00b050"))  # dark green
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=2, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Sig_V")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "000000"  # black
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="000000"))  # black
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=10, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Sig_H")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "0070c0"  # blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="0070c0"))  # blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 16.25
    chart.width = 14.45
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.7, h=0.25, w=0.2))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws5.add_chart(chart, "A18")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = ""  # empty string as filler title to match alignment of other charts
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = halfway_dept // 100 * 100
    chart.x_axis.scaling.orientation = "maxMin"
    chart.x_axis.scaling.min = -100
    chart.x_axis.scaling.max = 50
    chart.y_axis.crosses = "max"
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws5, min_col=13, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Induce FRAC")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "0070c0"  # blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="0070c0"))  # blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=12, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Breakout")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "c8500a"  # brown
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="c8500a"))  # brown
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 16.25
    chart.width = 14.45
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.7, h=0.2, w=0.2))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws5.add_chart(chart, "H18")

    chart = ScatterChart()
    cp = CharacterProperties(sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title = "Ddepth (m)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.min = halfway_dept // 100 * 100
    chart.x_axis.title = "Gradient (kPa/m)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1400, b=False))
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 40
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    values = Reference(ws5, min_col=14, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Pore pressure")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "00b0f0"  # light blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="00b0f0"))  # light blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=16, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Sig_h")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "00b050"  # dark green
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="00b050"))  # dark green
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=15, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Sig_V")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "000000"  # black
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="000000"))  # black
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    values = Reference(ws5, min_col=17, min_row=13, max_row=i - halfway_len + 8)
    series = Series(x_values, values, title="Sig_H")
    series.marker.symbol = "circle"
    series.marker.graphicalProperties.solidFill = "0070c0"  # blue
    series.marker.graphicalProperties.line.noFill = True
    series.marker.size = 2
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="0070c0"))  # blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 0.75 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 16.25
    chart.width = 14.45
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.05, h=0.9, w=0.85))
    chart.legend.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.2, y=0.7, h=0.25, w=0.2))
    cp = CharacterProperties(sz=1200)
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    ws5.add_chart(chart, "P18")

    ws6 = wb.create_sheet("Crtitical state M&L")

    ws6.column_dimensions["A"].width = 13.78 + width_const
    ws6.column_dimensions["D"].width = 11.33 + width_const
    ws6.column_dimensions["E"].width = 10.11 + width_const
    ws6.column_dimensions["I"].width = 10.89 + width_const

    for row_count, row in enumerate(ws6["A1":"M4"]):
        for col_count, cell in enumerate(row):
            cell.value = [None, None, None, "total", "eff", "Pp", "Pp increase", None, "Formation", "TVD, m", "Pp_G",
                          "Sig_G", None,
                          "Sig_1_eff", "=E2-G2", "MPa", "=$J$2*L2/1000", "=D2-F2", "=$J$2*$K$2/1000", 0, None,
                          "M & L", "=(stress!G6+stress!G7)/2", "=stress!I6", "=MAX(stress!O6:Q6)",
                          "=MID(INDEX(stress!$O$2:$Q$2,MATCH(L2,stress!$O$6:$Q$6,0)),5,1)",  # new stress gradient logic
                          "Sig_2_eff", "=E3-G3", "MPa", "=$J$2*L3/1000", "=D3-F3", "=$J$2*$K$2/1000", 0, None,
                          None, None, None, "=MEDIAN(stress!O6:Q6)",
                          "=MID(INDEX(stress!$O$2:$Q$2,MATCH(L3,stress!$O$6:$Q$6,0)),5,1)",
                          "Sig_3_eff", "=E4-G4", "MPa", "=$J$2*L4/1000", "=D4-F4", "=$J$2*$K$2/1000", 0, None,
                          None, None, None, "=MIN(stress!O6:Q6)",
                          "=MID(INDEX(stress!$O$2:$Q$2,MATCH(L4,stress!$O$6:$Q$6,0)),5,1)"][row_count * 13 + col_count]

    ws6["A6"], ws6["B6"], ws6["A8"], ws6["B8"], ws6["C8"], ws6["E8"] = "Friction coeeficient", 0.6, "Cohesion", \
        "=$B$5*(1-SIN($B$6*PI()/180))/(2*COS($B$6*PI()/180))", "MPa", "=B8*0"

    for row_count, row in enumerate(ws6["A10":"H12"]):
        for col_count, cell in enumerate(row):
            cell.value = [None, "Sig_1_3", None, "Sig_1-2", None, "Sig_2-3", None, "envelope",
                          "phi", "Sig_nor", "Sig_tao", "Sig_nor", "Sig_tao", "Sig_nor", "Sig_tao", None,
                          "deg", "MPa", "MPa", "MPa", "MPa", "MPa", "MPa", "MPa"][row_count * 8 + col_count]

    for i_ws6 in range(0, 181):
        ws6.append(
            {"A": i_ws6,
             "B": "=($B$2+$B$4)/2-($B$2-$B$4)/2*COS(A" + str(i_ws6 + 13) + "/180*PI())",
             "C": "=($B$2-$B$4)/2*SIN(A" + str(i_ws6 + 13) + "/180*PI())",
             "D": "=($B$2+$B$3)/2-($B$2-$B$3)/2*COS(A" + str(i_ws6 + 13) + "/180*PI())",
             "E": "=($B$2-$B$3)/2*SIN(A" + str(i_ws6 + 13) + "/180*PI())",
             "F": "=($B$3+$B$4)/2-($B$3-$B$4)/2*COS(A" + str(i_ws6 + 13) + "/180*PI())",
             "G": "=($B$3-$B$4)/2*SIN(A" + str(i_ws6 + 13) + "/180*PI())",
             "H": "=$E$8+A" + str(i_ws6 + 13) + "*$B$6"})

    for row in range(0, ws6.max_row + 1):
        ws6.row_dimensions[row].height = 15  # adjust height of rows after all have been generated

    chart = ScatterChart()
    cp = CharacterProperties(sz=1600)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "Effective normal stress (MPa)"
    chart.x_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1800, b=False))
    chart.y_axis.title = "Shear stress (MPa)"
    chart.y_axis.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1800, b=False))
    chart.x_axis.scaling.max = 40
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 16
    chart.x_axis.minorGridlines = ChartLines()  # have to initialise minor gridlines
    chart.y_axis.minorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.x_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="d0cece"))
    chart.y_axis.minorGridlines.spPr = GraphicalProperties(ln=drawing.line.LineProperties(solidFill="f2f2f2"))
    x_values = Reference(ws6, min_col=1, min_row=13, max_row=13+180)
    values = Reference(ws6, min_col=8, min_row=13, max_row=13+180)
    series = Series(values, x_values, title="Envilope")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="ff0000"))  # RED
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 1.5 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    x_values = Reference(ws6, min_col=2, min_row=13, max_row=13+180)
    values = Reference(ws6, min_col=3, min_row=13, max_row=13+180)
    series = Series(values, x_values, title="MC13")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="0070c0"))  # blue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 1.25 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    x_values = Reference(ws6, min_col=4, min_row=13, max_row=13+180)
    values = Reference(ws6, min_col=5, min_row=13, max_row=13+180)
    series = Series(values, x_values, title="MC12")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="00b0f0"))  # lightblue
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 1.25 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    x_values = Reference(ws6, min_col=6, min_row=13, max_row=13+180)
    values = Reference(ws6, min_col=7, min_row=13, max_row=13+180)
    series = Series(values, x_values, title="MC23")
    lineProp = drawing.line.LineProperties(solidFill=drawing.colors.ColorChoice(srgbClr="00b050"))  # green
    series.graphicalProperties.line = lineProp
    series.graphicalProperties.line.width = 1.25 * emu_to_pt_const
    series.smooth = True  # make the line smooth
    chart.series.append(series)
    chart.height = 11.06
    chart.width = 22.7
    chart.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.05, y=0.05, h=0.85, w=0.93))
    chart.legend = None
    ws6.add_chart(chart, "I12")

    return wb


def process_well_folder(folder_path):
    """Given a path to a north well information folder, return a dictionary of information from the folder files."""
    uwi, lati, long, kb, gl, mud_weight, formation_tops = None, None, None, None, None, [], []
    for file in os.listdir(folder_path):
        if file[-4:].lower() == ".las" and not mud_weight:  # alternatively, split basename to get file extension
            # only process if mud weight has not been found
            las = lasio.read(os.path.join(folder_path, file))
            # kb, gl = las.params.ekb.value, las.params.egl.value
            # if las.params.ekb.unit in {"F", "FT"}:  # in feet
            #     kb, gl = kb * 0.3048, gl * 0.3048
            try:
                mud_weight = [las.params.mudd.value, False]  # not assumed
                if las.params.mudd.unit in {"LBS/GAL", "lb/gal"}:  # in pounds per gallon
                    mud_weight[0] = mud_weight[0] * 119.8264273
            except AttributeError:  # clumsy approach
                try:
                    mud_weight = [las.params.dfd.value, False]  # second possible mnemonic
                    if las.params.dfd.unit in {"LBS/GAL", "lb/gal"}:  # in pounds per gallon
                        mud_weight[0] = mud_weight[0] * 119.8264273
                except AttributeError:  # missing mud weight in .las file
                    mud_weight = [1150, True]  # assumed
        elif file[-4:].lower() == ".txt":  # alternatively, split basename to get file extension
            file = open(os.path.join(folder_path, file))
            processing_formation = False
            for line in file:
                if "Well ID:" in line:
                    uwi = line.split()[2]
                elif "Govt KB" in line:
                    kb, gl = float(line.split()[2][:-1]), float(line.split()[4][:-1])
                elif "BH Lat:" in line:
                    lati = line.split()[-3].split(":")[1]  # split twice, first on space then on colon
                elif "BH Lon:" in line:
                    long = line.split()[-3].split(":")[1]  # split twice, first on space then on colon
                elif "Formation    TVD (m)    Elev (m)" in line:
                    processing_formation = True
                elif processing_formation:
                    formation = line.split()
                    if not formation or "~~~~~~~~~~~~~~~~~~~~~~~~~~~~" in formation[0]:  # newline or tildes denote end
                        break
                    else:
                        formation_set = {"trbaldnnl", "trhalfway", "trdoig",
                                         "trmntny_u", "trmontney", "trmntny_m", "trmntny_l", "prbelloy"}
                        #  inconsistent capitalisation in file/reference formation names, so checking lowercase
                        if formation[0].lower() in formation_set:
                            formation_tops.append(formation[:3])
                        if len(formation) > 3:  # two columns
                            if formation[3].lower() in formation_set:
                                formation_tops.append(formation[3:6])
                            if formation[4].lower() in formation_set:  # N Pay is present
                                formation_tops.append(formation[4:7])

    if uwi == "100/14-06-085-18W6/00":  # hardcoded special case
        formation_tops.append(["TRmntny_U", 1547, np.NaN])
        formation_tops.append(["TRmntny_M", 1634, np.NaN])
        formation_tops.append(["TRmntny_L", 1782, np.NaN])

    df_cols = ["Formation tops", "TVD", "Elev"]
    formation_tops = pd.DataFrame(formation_tops, columns=df_cols)
    formation_tops = formation_tops.replace({"TRmontney": "Trmontney (TRmntny_M)",
                                             "TRmntny_M": "Trmontney (TRmntny_M)"})

    if "TRmntny_U" not in formation_tops["Formation tops"].values:
        index = formation_tops[formation_tops["Formation tops"] == "Trmontney (TRmntny_M)"].index[0]
        formation_tops = pd.concat([pd.DataFrame([["TRmntny_U",
                                                   formation_tops.iloc[index]["TVD"],
                                                   formation_tops.iloc[index]["Elev"]]],
                                                 columns=df_cols),
                                    formation_tops]).reset_index(drop=True)  # may be overcomplicated

    if "TRmntny_L" not in formation_tops["Formation tops"].values:
        index = formation_tops[formation_tops["Formation tops"] == "PRbelloy"].index[0]
        formation_tops = pd.concat([pd.DataFrame([["TRmntny_L",
                                                   formation_tops.iloc[index]["TVD"],
                                                   formation_tops.iloc[index]["Elev"]]],
                                                 columns=df_cols),
                                    formation_tops]).reset_index(drop=True)  # may be overcomplicated

    return {"uwi": uwi, "lati": lati, "long": long, "kb": kb, "gl": gl,
            "mud_weight": mud_weight, "formation_tops": formation_tops}


def main():
    # noinspection PyUnusedLocal
    app = QApplication(sys.argv)  # just to keep QApplication in memory, a gui event loop with exec_() isn't needed
    print("Select the folder containing all north well folders to process basic information from")
    north_well_path = QFileDialog.getExistingDirectory()
    if not north_well_path:  # cancelled
        sys.exit()

    print("Select the file containing linear coefficients to predict DTSM using DTCO and Depth")
    correlation_path = QFileDialog.getOpenFileName()[0]
    if not correlation_path:  # cancelled
        sys.exit()
    coefficient_df = pd.read_csv(correlation_path)

    root, folder_names = next(os.walk(north_well_path))[:2]
    folder_paths = [os.path.join(root, folder) for folder in folder_names]
    for i, folder_path in enumerate(folder_paths):
        info = process_well_folder(folder_path)

        print()
        for key in info:
            if key != "formation_tops":
                print(key + ": \t" + str(info[key]))
            else:
                print(info[key].to_string(index=False))
        # nicer printing

        las_df = pd.read_csv(os.path.join(root, folder_names[i] + ".csv"))
        wb = make_excel_manual(info, las_df, coefficient_df)
        # wb = make_excel_copy(info, reference_file_path)  # too slow and does not copy styles properly
        wb.save(os.path.join(root, folder_names[i] + " geomechanics.xlsx"))


if __name__ == "__main__":
    my_module = modify_and_import("openpyxl.chart.chartspace", None,
                                  lambda src: src.replace("autoTitleDeleted=None", "autoTitleDeleted=True"))
    # workaround to not being able to remove title space through any regular method in the library
    # source code in https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/chart/chartspace.html#ChartContainer
    main()
